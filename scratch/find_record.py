import openpyxl

wb = openpyxl.load_workbook('/Users/akarsu/Desktop/2026-04.xlsx', read_only=True)
ws = wb.active

# Let's find headers
headers = []
for row in ws.iter_rows(min_row=1, max_row=10):
    vals = [cell.value for cell in row]
    if any(vals):
        # Check if this row looks like header
        if any(v in ['Tesisat', 'TESİSAT', 'Tesisat No', 'TESİSAT NO'] for v in vals if v):
            headers = [str(v).strip() if v is not None else '' for v in vals]
            header_row_index = row[0].row
            print(f"Header found at row {header_row_index}")
            break

if not headers:
    # Fallback to first row
    for row in ws.iter_rows(min_row=1, max_row=1):
        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in row]
        header_row_index = 1
    print("Fallback to row 1 as header")

# Print all columns that contain 'birim' or 'fiyat' (case-insensitive)
for idx, h in enumerate(headers):
    if 'birim' in h.lower() or 'fiyat' in h.lower() or 'fıyar' in h.lower() or 'fıyat' in h.lower():
        print(f"Col {idx}: '{h}'")

# Find the row for tesisat 4914999
tesisat_col_idx = -1
for idx, h in enumerate(headers):
    if h.lower() in ['tesisat', 'tesi̇sat', 'tesisat no', 'tesi̇sat no', 'abone_tesis_no']:
        tesisat_col_idx = idx
        break

if tesisat_col_idx == -1:
    # search for 'tesisat' in header names
    for idx, h in enumerate(headers):
        if 'tesisat' in h.lower() or 'tesi̇sat' in h.lower():
            tesisat_col_idx = idx
            break

print(f"Tesisat column index: {tesisat_col_idx} (Name: '{headers[tesisat_col_idx]}')")

# Search rows
for row in ws.iter_rows(min_row=header_row_index + 1):
    vals = [cell.value for cell in row]
    if not any(vals):
        continue
    if len(vals) > tesisat_col_idx and vals[tesisat_col_idx] is not None:
        val_str = str(vals[tesisat_col_idx]).strip()
        if val_str == '4914999':
            print("Found record row:")
            for idx, (h, v) in enumerate(zip(headers, vals)):
                print(f"[{idx}] {h}: {v}")
            break
