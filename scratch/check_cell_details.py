import openpyxl

wb = openpyxl.load_workbook('/Users/akarsu/Desktop/2026-04.xlsx', data_only=False)
ws = wb.active

# Let's find header
headers = []
for row in ws.iter_rows(min_row=1, max_row=1):
    headers = [cell.value for cell in row]

tesisat_col_idx = 6 # We found it is index 6 previously

for r_idx, row in enumerate(ws.iter_rows(min_row=2)):
    tesisat_val = row[tesisat_col_idx].value
    if tesisat_val is not None and str(tesisat_val).strip() == '4914999':
        print(f"Row {r_idx + 2} found.")
        for col_idx in [95, 96, 97, 98, 99]:
            cell = row[col_idx]
            print(f"Col {col_idx} ({headers[col_idx]}):")
            print(f"  Coordinate: {cell.coordinate}")
            print(f"  Raw Value (data_only=False): {cell.value}")
            
# Now load with data_only=True to see evaluated value
wb_data = openpyxl.load_workbook('/Users/akarsu/Desktop/2026-04.xlsx', data_only=True)
ws_data = wb_data.active
for r_idx, row in enumerate(ws_data.iter_rows(min_row=2)):
    tesisat_val = row[tesisat_col_idx].value
    if tesisat_val is not None and str(tesisat_val).strip() == '4914999':
        for col_idx in [95, 96, 97, 98, 99]:
            cell = row[col_idx]
            print(f"Col {col_idx} Evaluated Value (data_only=True): {cell.value}")
