#!/usr/bin/env python3
"""Excel'i CSV'ye dönüştürür"""
import openpyxl, csv, sys

src = '/Users/akarsu/Desktop/kuyular.xlsx'
dst = '/Users/akarsu/Desktop/suski/storage/app/kuyular_import.csv'

wb = openpyxl.load_workbook(src)
ws = wb.active

with open(dst, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f, quoting=csv.QUOTE_ALL)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(['' if v is None else str(v) for v in row])

rows = ws.max_row - 1
print(f'✅ CSV oluşturuldu: {dst}  ({rows} veri satırı)')
